"""Render an OPD eval ``OUTPUT_ROOT`` into a **methods × benchmarks** report.

Rows = methods (the ``MODELS`` tags), columns = benchmarks; every metric shown as a
**percentage** (e.g. ``45.23%``), not a ``0.x`` ratio. Reads exactly the
``summary.json`` files ``scripts/eval_opd_multi.sh`` writes — both shapes:

* judged group (``run_opd_eval``): ``{"model_name", "datasets"/"benchmarks": [{"dataset", "pass_at_k"}]}``
* deterministic group (``run_vqa_eval``): ``{"model_name", "benchmarks": {name: {"metrics": {...}}}}``
  (POPE F1 / ChartQA relaxed acc / VQAv2 soft acc).

Writes, under ``OUTPUT_ROOT`` (or ``--out-dir``):
  * ``report.md``  — Markdown table (always)
  * ``report.csv`` — same table, opens directly in Excel (always)
  * ``report.xlsx``— real spreadsheet with ``0.00%`` cell formatting (only if
    ``openpyxl`` is importable; otherwise skipped with a hint)
and prints the Markdown table to stdout.

This is a pure-stdlib transpose/format of the matrix the inline aggregation in
``eval_opd_multi.sh`` already computes, so it pulls in no torch/vLLM and can be run
any time on a finished (or partially finished) ``OUTPUT_ROOT``:

    python3 baseline/eval/make_report.py eval_outputs/626Newprompt
    # or, equivalently:
    python -m baseline.eval.make_report eval_outputs/626Newprompt --decimals 1
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import os

# Deterministic benchmark name -> (metric key in summary['benchmarks'][name]['metrics'],
# column label). Mirrors the inline aggregation in scripts/eval_opd_multi.sh.
DET_METRIC = {
    "pope": ("f1", "pope (F1)"),
    "chartqa": ("relaxed_accuracy", "chartqa (relax)"),
    "vqav2": ("vqa_accuracy", "vqav2 (soft)"),
    "vstar": ("accuracy", "vstar (acc)"),
}

# Derived column: the average of MMMU-Pro's two sub-scores. It re-aggregates columns
# already in the table, so it is kept OUT of the judged average (registered in
# det_labels, which judged_avg() skips). Emitted only when ≥1 sub-score actually ran.
PRO_SUBSCORES = ("mmmu_pro_10options", "mmmu-pro-vision")
PRO_AVG_LABEL = "mmmu-pro (avg)"

# Preferred column order (the standard suite); unknown benchmarks sort alphabetically
# after these.
BENCH_ORDER = [
    "mathvista", "MathVista", "mathverse", "MathVerse", "mathvision", "MathVision",
    "MMMU", "mmmu_pro_10options", "mmmu-pro-vision", PRO_AVG_LABEL, "MMMU-Pro",
    "mmstar", "MMStar", "hallusionbench", "HallusionBench",
    "pope (F1)", "POPE", "chartqa (relax)", "ChartQA", "vqav2 (soft)",
    "vstar (acc)", "V* Bench", "HRBench4K", "HRBench8K", "MME-RealWorld-Lite",
]

AVG_LABEL = "Avg (judged)"


def collect(root: str):
    """Parse every ``summary.json`` under ``root`` into ``matrix[benchmark][tag] = ratio``.

    Returns ``(matrix, tags, det_labels)`` where ``det_labels`` is the set of
    benchmark columns kept out of the judged average — either scored by an official
    metric or a derived re-aggregation (the MMMU-Pro sub-score average).
    """
    matrix: dict[str, dict[str, float | None]] = {}
    tags: list[str] = []
    det_labels: set[str] = set()
    for path in sorted(glob.glob(os.path.join(root, "**", "summary.json"), recursive=True)):
        try:
            summary = json.load(open(path, encoding="utf-8"))
        except Exception:
            continue
        # model_name is written by both evals; fall back to the folder name (the tag
        # dir is one level above the per-dataset dir).
        tag = summary.get("model_name") or os.path.relpath(
            os.path.dirname(os.path.dirname(path)), root
        )
        if tag not in tags:
            tags.append(tag)
        bms = summary.get("benchmarks")
        if summary.get("backend") in {"lmms-eval", "lmms-eval-fast"} and isinstance(bms, dict):
            for name, score in bms.items():
                label = (score or {}).get("label") or name
                val = (score or {}).get("score")
                if not isinstance(val, (int, float)):
                    metric = (score or {}).get("primary_metric")
                    metrics = (score or {}).get("metrics") or {}
                    val = metrics.get(metric) if metric else None
                matrix.setdefault(label, {})[tag] = val if isinstance(val, (int, float)) else None
                det_labels.add(label)
        elif isinstance(bms, dict):  # deterministic (run_vqa_eval): name -> {metrics: {...}}
            for name, score in bms.items():
                key, label = DET_METRIC.get(name, (None, name))
                val = ((score or {}).get("metrics") or {}).get(key) if key else None
                matrix.setdefault(label, {})[tag] = val
                det_labels.add(label)
        else:  # judged (run_opd_eval): datasets/benchmarks lists -> pass_at_k
            for entry in (summary.get("datasets", []) + (bms or [])):
                name = os.path.basename(
                    str(entry.get("dataset") or entry.get("benchmark") or "?").rstrip("/")
                )
                matrix.setdefault(name, {})[tag] = entry.get("pass_at_k")

    # Derived MMMU-Pro average column = per-method mean of the two sub-scores. Added to
    # det_labels so it stays out of the judged average (it only re-aggregates columns
    # already counted there). Skipped entirely if neither sub-score was evaluated.
    if any(sub in matrix for sub in PRO_SUBSCORES):
        avg_col: dict[str, float | None] = {}
        for tag in tags:
            sub_vals = [
                matrix[sub][tag]
                for sub in PRO_SUBSCORES
                if isinstance(matrix.get(sub, {}).get(tag), (int, float))
            ]
            avg_col[tag] = sum(sub_vals) / len(sub_vals) if sub_vals else None
        matrix[PRO_AVG_LABEL] = avg_col
        det_labels.add(PRO_AVG_LABEL)
    return matrix, tags, det_labels


def order_benchmarks(names) -> list[str]:
    rank = {n: i for i, n in enumerate(BENCH_ORDER)}
    return sorted(names, key=lambda n: (rank.get(n, len(BENCH_ORDER)), n))


def judged_avg(matrix, benches, det_labels, tag):
    vals = [
        matrix[b][tag]
        for b in benches
        if b not in det_labels and isinstance(matrix[b].get(tag), (int, float))
    ]
    return sum(vals) / len(vals) if vals else None


def pct(value, decimals: int) -> str:
    return f"{value * 100:.{decimals}f}%" if isinstance(value, (int, float)) else "-"


def build_rows(matrix, tags, det_labels, decimals: int):
    """(header, rows) where each row is [method, *pct cells, avg pct]."""
    benches = order_benchmarks(matrix.keys())
    header = ["Method", *benches, AVG_LABEL]
    rows = []
    for tag in tags:
        cells = [pct(matrix[b].get(tag), decimals) for b in benches]
        avg = pct(judged_avg(matrix, benches, det_labels, tag), decimals)
        rows.append([tag, *cells, avg])
    return header, rows, benches


def write_markdown(path, header, rows):
    # Right-align every numeric column; left-align the Method column.
    align = [":---"] + ["---:"] * (len(header) - 1)
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(align) + " |",
        *["| " + " | ".join(r) + " |" for r in rows],
    ]
    text = "\n".join(lines) + "\n"
    open(path, "w", encoding="utf-8").write(text)
    return text


def write_csv(path, header, rows):
    with open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)


def write_xlsx(path, header, rows, matrix, tags, benches, det_labels, decimals: int) -> bool:
    """Real spreadsheet: numeric ratios with a ``0.00%`` cell format (so they stay
    sortable/computable while displaying as percentages). Needs openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception:
        return False
    num_fmt = "0." + "0" * decimals + "%" if decimals > 0 else "0%"
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for tag in tags:
        row = [tag]
        for b in benches:
            v = matrix[b].get(tag)
            row.append(v if isinstance(v, (int, float)) else None)
        avg = judged_avg(matrix, benches, det_labels, tag)
        row.append(avg if isinstance(avg, (int, float)) else None)
        ws.append(row)
    # Percentage format on every numeric cell (everything but column A).
    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            ws.cell(row=r, column=c).number_format = num_fmt
    ws.freeze_panes = "B2"  # keep method names + header visible while scrolling
    ws.column_dimensions["A"].width = max((len(t) for t in tags), default=8) + 2
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(
            12, len(str(header[col - 1])) + 2
        )
    wb.save(path)
    return True


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("root", help="eval OUTPUT_ROOT (the dir holding <tag>/<dataset>/summary.json)")
    ap.add_argument(
        "--decimals", type=int, default=int(os.environ.get("REPORT_DECIMALS", "2")),
        help="percentage decimal places (default 2, e.g. 45.23%%)",
    )
    ap.add_argument(
        "--methods", default="",
        help="comma list to fix the method (row) order; default = discovery order, "
        "extras appended alphabetically",
    )
    ap.add_argument("--out-dir", default="", help="where to write the report (default = root)")
    args = ap.parse_args()

    if not os.path.isdir(args.root):
        raise SystemExit(f"OUTPUT_ROOT not found: {args.root}")
    matrix, tags, det_labels = collect(args.root)
    if not matrix:
        raise SystemExit(
            f"No summary.json found under {args.root} — has the eval finished "
            "(judge/all phase)? Nothing to report."
        )

    if args.methods:
        want = [m.strip() for m in args.methods.split(",") if m.strip()]
        tags = [m for m in want if m in tags] + [t for t in sorted(tags) if t not in want]
    else:
        tags = sorted(tags)

    header, rows, benches = build_rows(matrix, tags, det_labels, args.decimals)
    out_dir = args.out_dir or args.root
    os.makedirs(out_dir, exist_ok=True)
    md_path = os.path.join(out_dir, "report.md")
    csv_path = os.path.join(out_dir, "report.csv")
    xlsx_path = os.path.join(out_dir, "report.xlsx")

    md = write_markdown(md_path, header, rows)
    write_csv(csv_path, header, rows)
    has_xlsx = write_xlsx(
        xlsx_path, header, rows, matrix, tags, benches, det_labels, args.decimals
    )

    print(md)
    print(f"Wrote {md_path}")
    print(f"Wrote {csv_path}")
    if has_xlsx:
        print(f"Wrote {xlsx_path}")
    else:
        print("(report.xlsx skipped: openpyxl not installed — `uv pip install openpyxl` "
              "for a formatted .xlsx; report.csv already opens in Excel.)")
    print(
        "\nColumns = benchmarks, rows = methods; metrics in %. Judged columns = pass@k "
        "(LLM judge); pope/chartqa/vqav2 = official metric. "
        f"'{AVG_LABEL}' averages the judged columns only."
    )


if __name__ == "__main__":
    main()
